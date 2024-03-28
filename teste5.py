import base64
import io
import urllib.request
from openpyxl import load_workbook
import streamlit as st
import pandas as pd

st.set_option('deprecation.showfileUploaderEncoding', False)


def create_onedrive_directdownload(onedrive_link):
    data_bytes64 = base64.b64encode(bytes(onedrive_link, 'utf-8'))
    data_bytes64_String = data_bytes64.decode('utf-8').replace('/', '_').replace('+', '-').rstrip("=")
    resultUrl = f"https://api.onedrive.com/v1.0/shares/u!{data_bytes64_String}/root/content"
    return resultUrl


### Original link copied from the file in onedrive
#onedrive_link = "https://1drv.ms/x/s!AnABowEQ_9D-gWfZKKiRfuYrOO5u?e=CaPd6q"
#onedrive_link = "https://1drv.ms/x/s!AnABowEQ_9D-gVqQMmcBZpJHvxIj?e=MIrs7F"
onedrive_link = "https://1drv.ms/x/s!AnABowEQ_9D-gWXHpo6enAouzQHi"
### Converted Onedrive link


import random

onedrive_direct_link = f"{create_onedrive_directdownload(onedrive_link)}&random={random.random()}"

onedrive_direct_link = create_onedrive_directdownload(onedrive_link)
### Retrieve url as bytes file
file = urllib.request.urlopen(onedrive_direct_link).read()

### Load file into Openpyxl
wb = load_workbook(filename=io.BytesIO(file))
#ws = wb['Planilha1']
#ws = wb['2024.1']
ws = wb['Alunos - Presença nas aulas']

### Transform ws into a DataFrame
#data = []
#for row in ws.iter_rows():
#    data.append({ws.cell(row=1, column=col).value: value for col, value in enumerate(row, start=1)})

# Transform ws into a DataFrame
data = []
for row in ws.iter_rows(values_only=True):
    data.append({f'Column_{col_num}': value for col_num, value in enumerate(row, start=1)})

df = pd.DataFrame(data)

### Convert 'Pessoas' column to string
#df['Pessoas'] = df['Pessoas'].astype(str)
df = df.astype(str)

### Show DataFrame using Streamlit
st.title("Data from OneDrive Worksheet")
st.table(df)