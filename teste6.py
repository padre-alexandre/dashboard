import base64
import io
import urllib.request
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import random

# Desativar o cache no Streamlit
st.set_option('deprecation.showfileUploaderEncoding', False)

def create_onedrive_directdownload(onedrive_link):
    data_bytes64 = base64.b64encode(bytes(onedrive_link, 'utf-8'))
    data_bytes64_String = data_bytes64.decode('utf-8').replace('/', '_').replace('+', '-').rstrip("=")
    resultUrl = f"https://api.onedrive.com/v1.0/shares/u!{data_bytes64_String}/root/content"
    return resultUrl

@st.cache(ttl=5, max_entries=1)
def load_data(onedrive_link, sheet_name):
    onedrive_direct_link = f"{create_onedrive_directdownload(onedrive_link)}&random={random.random()}"
    file = urllib.request.urlopen(onedrive_direct_link).read()
    wb = load_workbook(filename=io.BytesIO(file))
    
    # Utilize o nome da aba desejada (no caso, '2024.1')
    ws = wb[sheet_name]
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append({ws.cell(row=1, column=col).value: value for col, value in enumerate(row, start=1)})
    df = pd.DataFrame(data)
    return df

# Novo link copiado da planilha no onedrive (com aba 2024.1)
onedrive_link = "https://1drv.ms/x/s!AnABowEQ_9D-gVqQMmcBZpJHvxIj?e=MIrs7F"
# Nome da aba desejada
sheet_name = "2024.1"

# Intervalo de atualização
#st.experimental_refresh(interval=5 * 1000)  # 5 segundos em milissegundos

# Carregar os dados
df = load_data(onedrive_link, sheet_name)

# Mostrar o DataFrame usando Streamlit
st.title("Dados da planilha do OneDrive")
st.dataframe(df)
